"""
Core analysis logic for the MetrixND Binary Tuning app.

Reads a MetrixND xlsx export (sheets: Data, Coef, MStat, Err — DStat/Corr/Elas/BX/YHat
are ignored) and produces a ranked list of binary-variable recommendations:
which existing binaries look safe to drop, which missing calendar months look
worth adding, and which individual data points look like data-quality issues
rather than real seasonal effects.

IMPORTANT CAVEAT: MetrixND's real fit can include AR(1)/MA(1) autocorrelation
correction. This module has no access to that — it fits a plain OLS proxy on
whatever columns are present in the Data sheet. For models with a
near-unit-root AR(1) term (AR(1) close to 1), the proxy can badly overstate
the importance of a candidate binary, because the AR term already absorbs a
persistent level shift that a static OLS model has no way to explain except
through a dummy variable. Always treat proxy significance as a lead to verify
in MetrixND, not a final answer. See `AR_CAVEAT` for the exact wording shown
in the app.
"""

from __future__ import annotations

import math
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Optional

import numpy as np
import openpyxl
import pandas as pd

MONTH_ABBR = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

AR_CAVEAT = (
    "These AIC/BIC/p-values come from a plain OLS proxy fit in this app, "
    "not from MetrixND itself. MetrixND's real model may include AR(1)/MA(1) "
    "autocorrelation terms this proxy cannot replicate. When AR(1) is close "
    "to 1 (near unit-root), the real model can already absorb a persistent "
    "shift that this proxy will misattribute to a candidate binary — always "
    "re-fit any accepted suggestion in MetrixND itself before trusting it."
)

REQUIRED_SHEETS = ["Data", "Coef", "MStat", "Err"]


class ParseError(Exception):
    pass


def _normal_sf_two_sided(z: np.ndarray) -> np.ndarray:
    """Two-sided p-value from the standard normal distribution, used as a
    large-sample approximation to the t-distribution (avoids a scipy
    dependency; with the sample sizes these models use — generally 100+
    monthly observations — this tracks the exact t-distribution p-value
    closely)."""
    z = np.abs(z)
    erf_vec = np.vectorize(math.erf)
    cdf = 0.5 * (1.0 + erf_vec(z / math.sqrt(2)))
    return 2 * (1 - cdf)


# --------------------------------------------------------------------------
# Header / column helpers
# --------------------------------------------------------------------------

def normalize_header(value) -> str:
    """MetrixND/Excel often stores one-off event dummy headers (e.g. "Jan-18"
    meaning January 2018) as datetime cells once opened in Excel, with the
    real month in `.month` and the 2-digit year tucked into `.day`
    (e.g. datetime(2026, 1, 18) -> "Jan-18" -> label "Jan18"). Reconstruct a
    readable label for those; pass through plain strings untouched."""
    if isinstance(value, (datetime, date)):
        month = value.month
        yy = value.day
        if 1 <= month <= 12 and 1 <= yy <= 99:
            return f"{MONTH_ABBR[month - 1]}{yy:02d}"
        return str(value)
    if value is None:
        return ""
    return str(value).strip()


def month_number_from_name(name: str) -> Optional[int]:
    """Return 1-12 if `name` is exactly a bare month name (Jan/Feb/...),
    case-insensitive, with no trailing year digits. Returns None otherwise
    (e.g. "Jan18" is an event dummy, not a bare seasonal month dummy)."""
    n = name.strip().lower()
    for i, abbr in enumerate(MONTH_ABBR, start=1):
        if n == abbr.lower() or n == abbr.lower() + ".":
            return i
    full_names = ["january", "february", "march", "april", "may", "june",
                  "july", "august", "september", "october", "november", "december"]
    if n in full_names:
        return full_names.index(n) + 1
    return None


# --------------------------------------------------------------------------
# Data containers
# --------------------------------------------------------------------------

@dataclass
class ParsedWorkbook:
    file_name: str
    target_name: str
    data_df: pd.DataFrame          # actual observations, regressor columns only + target
    regressor_cols: list            # all regressor column names (order preserved)
    month_dummy_cols: dict          # month number -> column name, for bare seasonal dummies
    missing_months: list             # month numbers with no bare seasonal dummy
    coef_df: pd.DataFrame           # as reported by MetrixND (Coef sheet)
    mstat: dict                     # as reported by MetrixND (MStat sheet)
    err_df: pd.DataFrame            # residuals (Err sheet)


def _sheet_rows(ws):
    return list(ws.iter_rows(values_only=True))


def load_workbook(file) -> ParsedWorkbook:
    wb = openpyxl.load_workbook(file, data_only=True)
    missing = [s for s in REQUIRED_SHEETS if s not in wb.sheetnames]
    if missing:
        raise ParseError(
            f"This doesn't look like a MetrixND export — missing sheet(s): {', '.join(missing)}. "
            f"Sheets found: {', '.join(wb.sheetnames)}"
        )

    data_df, target_name, regressor_cols, month_dummy_cols, missing_months = _parse_data_sheet(
        wb["Data"]
    )
    coef_df = _parse_coef_sheet(wb["Coef"])
    mstat = _parse_mstat_sheet(wb["MStat"])
    err_df = _parse_err_sheet(wb["Err"])

    return ParsedWorkbook(
        file_name=getattr(file, "name", "uploaded_file.xlsx"),
        target_name=target_name,
        data_df=data_df,
        regressor_cols=regressor_cols,
        month_dummy_cols=month_dummy_cols,
        missing_months=missing_months,
        coef_df=coef_df,
        mstat=mstat,
        err_df=err_df,
    )


def _parse_data_sheet(ws):
    rows = _sheet_rows(ws)
    if len(rows) < 2:
        raise ParseError("Data sheet is empty.")
    header_raw = rows[0]
    headers = [normalize_header(h) for h in header_raw]
    if len(headers) < 3:
        raise ParseError("Data sheet needs at least Year, Month, Target columns.")

    target_name = headers[2]
    exclude = {headers[0], headers[1], target_name, "XMissing", "YMissing", ""}

    seen = set()
    regressor_cols = []
    for h in headers[3:]:
        if h in exclude or h in seen:
            continue
        seen.add(h)
        regressor_cols.append(h)

    body = [r for r in rows[1:] if r[2] is not None]
    if not body:
        raise ParseError("No data rows found (target column is empty).")

    cols = ["Year", "Month", target_name] + headers[3:]
    df = pd.DataFrame(body, columns=cols[: len(body[0])])
    keep_cols = ["Year", "Month", target_name] + regressor_cols
    keep_cols = [c for c in keep_cols if c in df.columns]
    df = df[keep_cols].copy()
    for c in df.columns:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    df = df.dropna(subset=[target_name])

    month_dummy_cols = {}
    for h in regressor_cols:
        mnum = month_number_from_name(h)
        if mnum is not None:
            month_dummy_cols[mnum] = h
    missing_months = [mm for mm in range(1, 13) if mm not in month_dummy_cols]

    return df, target_name, regressor_cols, month_dummy_cols, missing_months


def _parse_coef_sheet(ws) -> pd.DataFrame:
    rows = _sheet_rows(ws)
    if not rows:
        return pd.DataFrame(columns=["Variable", "Coefficient", "StdErr", "TStat", "PValue", "Definition"])
    header = [normalize_header(h) for h in rows[0]]
    body = [r for r in rows[1:] if r and r[0] is not None]
    n_cols = len(header)
    clean_rows = [list(r[:n_cols]) + [None] * (n_cols - len(r)) for r in body]
    df = pd.DataFrame(clean_rows, columns=header)
    rename = {}
    for c in df.columns:
        cl = c.lower()
        if cl.startswith("variable"):
            rename[c] = "Variable"
        elif cl.startswith("coefficient"):
            rename[c] = "Coefficient"
        elif "std" in cl and "err" in cl:
            rename[c] = "StdErr"
        elif "t-stat" in cl or cl == "tstat":
            rename[c] = "TStat"
        elif "p-value" in cl or cl == "pvalue":
            rename[c] = "PValue"
        elif cl.startswith("definition"):
            rename[c] = "Definition"
    df = df.rename(columns=rename)
    for c in ["Coefficient", "StdErr", "TStat", "PValue"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    return df


def _parse_mstat_sheet(ws) -> dict:
    rows = _sheet_rows(ws)
    stats_dict = {}
    label_map = {
        "r-squared": "R2",
        "adjusted r-squared": "AdjR2",
        "aic": "AIC",
        "bic": "BIC",
        "durbin-watson statistic": "DW",
        "ljung-box statistic": "LjungBox",
        "prob (ljung-box)": "LjungBox_p",
        "skewness": "Skewness",
        "kurtosis": "Kurtosis",
        "jarque-bera": "JarqueBera",
        "prob (jarque-bera)": "JarqueBera_p",
        "mean abs. % err. (mape)": "MAPE",
        "adjusted observations": "N",
        "forecast observations": "ForecastN",
    }
    for row in rows:
        if not row:
            continue
        label = normalize_header(row[0]).strip().lower()
        if label in label_map and len(row) > 1:
            stats_dict[label_map[label]] = row[1]
        if len(row) > 3 and row[3] is not None:
            flabel = normalize_header(row[3]).strip().lower()
            if flabel == "mean abs. % err. (mape)" and len(row) > 4:
                stats_dict["ForecastMAPE"] = row[4]
    return stats_dict


def _parse_err_sheet(ws) -> pd.DataFrame:
    rows = _sheet_rows(ws)
    if not rows:
        return pd.DataFrame(columns=["Year", "Month", "Actual", "Pred", "Resid", "PctResid", "StdResid"])
    header = [normalize_header(h) for h in rows[0]]
    body = [r for r in rows[1:] if r and len(r) > 4 and r[4] is not None]
    n_cols = min(len(header), 7)
    clean_rows = [list(r[:n_cols]) for r in body]
    cols = ["Year", "Month", "Actual", "Pred", "Resid", "PctResid", "StdResid"][:n_cols]
    df = pd.DataFrame(clean_rows, columns=cols)
    for c in df.columns:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    return df


# --------------------------------------------------------------------------
# OLS proxy
# --------------------------------------------------------------------------

@dataclass
class FitResult:
    cols: list
    k: int
    n: int
    r2: float
    adj_r2: float
    aic: float
    bic: float
    mape: float
    coefs: np.ndarray = field(repr=False)
    se: np.ndarray = field(repr=False)
    tstat: np.ndarray = field(repr=False)
    pvalue: np.ndarray = field(repr=False)


def ols_fit(df: pd.DataFrame, target: str, cols: list) -> FitResult:
    cols = [c for c in cols if c in df.columns]
    y = df[target].to_numpy(dtype=float)
    Xraw = df[cols].to_numpy(dtype=float) if cols else np.zeros((len(df), 0))
    X = np.column_stack([np.ones(len(df)), Xraw])
    beta, _, rank, _ = np.linalg.lstsq(X, y, rcond=None)
    resid = y - X @ beta
    n, k = X.shape
    dof = max(n - k, 1)
    sse = float(np.sum(resid ** 2))
    sst = float(np.sum((y - y.mean()) ** 2)) or 1e-9
    r2 = 1 - sse / sst
    adj_r2 = 1 - (1 - r2) * (n - 1) / dof
    safe_sse_n = max(sse / n, 1e-12)
    aic = n * np.log(safe_sse_n) + 2 * k
    bic = n * np.log(safe_sse_n) + k * np.log(n)
    nonzero_y = np.where(y != 0, y, np.nan)
    mape = float(np.nanmean(np.abs(resid / nonzero_y))) * 100

    sigma2 = sse / dof
    try:
        xtx_inv = np.linalg.pinv(X.T @ X)
        se = np.sqrt(np.clip(np.diag(xtx_inv) * sigma2, 0, None))
        tstat = np.divide(beta, se, out=np.zeros_like(beta), where=se > 0)
        pvalue = _normal_sf_two_sided(tstat)
    except np.linalg.LinAlgError:
        se = np.full_like(beta, np.nan)
        tstat = np.full_like(beta, np.nan)
        pvalue = np.full_like(beta, np.nan)

    return FitResult(
        cols=cols, k=k, n=n, r2=r2, adj_r2=adj_r2, aic=aic, bic=bic, mape=mape,
        coefs=beta, se=se, tstat=tstat, pvalue=pvalue,
    )


# --------------------------------------------------------------------------
# By-month bias + outliers
# --------------------------------------------------------------------------

def by_month_bias(err_df: pd.DataFrame) -> pd.DataFrame:
    if err_df.empty:
        return pd.DataFrame(columns=["Month", "MonthName", "N", "MeanResid"])
    g = err_df.groupby("Month")["Resid"].agg(["count", "mean"]).reset_index()
    g.columns = ["Month", "N", "MeanResid"]
    g["MonthName"] = g["Month"].apply(lambda mm: MONTH_ABBR[int(mm) - 1] if 1 <= mm <= 12 else str(mm))
    return g.sort_values("Month")[["Month", "MonthName", "N", "MeanResid"]]


def find_outliers(err_df: pd.DataFrame, z_thresh: float = 2.5, top_n: int = 8) -> pd.DataFrame:
    if err_df.empty or "StdResid" not in err_df.columns:
        return pd.DataFrame(columns=["Year", "Month", "MonthName", "Resid", "PctResid", "StdResid"])
    df = err_df.copy()
    df["AbsStd"] = df["StdResid"].abs()
    flagged = df[df["AbsStd"] >= z_thresh].sort_values("AbsStd", ascending=False)
    if flagged.empty:
        flagged = df.sort_values("AbsStd", ascending=False).head(top_n)
    else:
        flagged = flagged.head(top_n)
    flagged = flagged.copy()
    flagged["MonthName"] = flagged["Month"].apply(lambda mm: MONTH_ABBR[int(mm) - 1] if 1 <= mm <= 12 else str(mm))
    return flagged[["Year", "Month", "MonthName", "Resid", "PctResid", "StdResid"]]


# --------------------------------------------------------------------------
# Recommendation engine
# --------------------------------------------------------------------------

@dataclass
class Recommendation:
    rank: int
    category: str
    variable: str
    add_yes_no: str
    delta_aic: Optional[float]
    delta_bic: Optional[float]
    delta_mape: Optional[float]
    pvalue: Optional[float]
    rationale: str


def is_binary_column(df: pd.DataFrame, col: str) -> bool:
    vals = df[col].dropna().unique()
    return set(np.round(vals, 6)).issubset({0.0, 1.0})


def continuous_regressors(df: pd.DataFrame, regressor_cols: list) -> list:
    return [c for c in regressor_cols if c in df.columns and not is_binary_column(df, c)]


def sign_check(coef_df: pd.DataFrame, continuous_cols: list, p_thresh: float = 0.05) -> list:
    """Flag continuous (non-binary) regressors that come back with a
    significant NEGATIVE coefficient in the real MetrixND fit — often a sign
    of misspecification/collinearity rather than a genuine effect, though it
    can occasionally be a legitimate structural relationship."""
    flags = []
    if coef_df.empty or "Variable" not in coef_df.columns:
        return flags
    for col in continuous_cols:
        matches = coef_df[coef_df["Variable"].astype(str).str.contains(re.escape(col), case=False, na=False)]
        for _, row in matches.iterrows():
            coefv = row.get("Coefficient")
            pv = row.get("PValue")
            if coefv is not None and pv is not None and not pd.isna(coefv) and not pd.isna(pv):
                if coefv < 0 and pv < p_thresh:
                    flags.append(
                        f"{row['Variable']}: significant NEGATIVE coefficient ({coefv:.4f}, p={pv:.4f}). "
                        f"Verify this matches the expected physical/economic direction — could be a real "
                        f"effect, or a sign of collinearity with the seasonal dummies."
                    )
    return flags


def existing_binary_removal_candidates(coef_df: pd.DataFrame, p_thresh: float = 0.05) -> pd.DataFrame:
    """Flag currently-included binary/event variables (mBin.* in MetrixND
    naming, or anything not a continuous driver) that are not significant in
    the REAL reported model — these are read straight from the Coef sheet,
    not the OLS proxy, since that's the more trustworthy source for variables
    already in the model."""
    if coef_df.empty or "PValue" not in coef_df.columns:
        return pd.DataFrame(columns=["Variable", "Coefficient", "PValue"])
    cand = coef_df[
        coef_df["Variable"].astype(str).str.contains(r"\.", na=False)
        & ~coef_df["Variable"].astype(str).str.contains(r"XOther|XHeat|AR\(|MA\(", case=False, na=False)
    ].copy()
    cand = cand[cand["PValue"] > p_thresh]
    return cand[["Variable", "Coefficient", "PValue"]].sort_values("PValue", ascending=False)


def forward_select_months(
    df: pd.DataFrame,
    target: str,
    base_cols: list,
    missing_months: list,
    bic_pref: bool = True,
):
    """Greedy forward selection over missing calendar months. At each step,
    try adding every remaining candidate month dummy to the current column
    set; accept the one that most improves the chosen criterion (BIC by
    default, matching the parsimony rule used throughout manual analysis this
    session); stop when no remaining candidate improves it further. Returns
    the list of accepted month numbers, plus a step-by-step trace table."""
    trace_cols = ["Step", "Candidate", "AIC", "BIC", "MAPE", "Improves"]
    if not missing_months:
        return [], pd.DataFrame(columns=trace_cols)

    work_df = df.copy()
    for mm in missing_months:
        col = f"_cand_{MONTH_ABBR[mm - 1]}"
        work_df[col] = (work_df["Month"] == mm).astype(float)

    current_cols = list(base_cols)
    baseline = ols_fit(work_df, target, current_cols)
    accepted = []
    trace_rows = []
    remaining = list(missing_months)

    current_metric = baseline.bic if bic_pref else baseline.aic

    while remaining:
        best = None
        best_metric = current_metric
        for mm in remaining:
            col = f"_cand_{MONTH_ABBR[mm - 1]}"
            trial_cols = current_cols + [col]
            fit = ols_fit(work_df, target, trial_cols)
            metric = fit.bic if bic_pref else fit.aic
            trace_rows.append({
                "Step": len(accepted) + 1,
                "Candidate": MONTH_ABBR[mm - 1],
                "AIC": round(fit.aic, 2),
                "BIC": round(fit.bic, 2),
                "MAPE": round(fit.mape, 3),
                "Improves": metric < current_metric,
            })
            if metric < best_metric:
                best_metric = metric
                best = mm
        if best is None:
            break
        accepted.append(best)
        current_cols.append(f"_cand_{MONTH_ABBR[best - 1]}")
        current_metric = best_metric
        remaining.remove(best)

    trace_df = pd.DataFrame(trace_rows, columns=trace_cols) if trace_rows else pd.DataFrame(columns=trace_cols)
    return accepted, trace_df


SOLO_DELTA_COLUMNS = ["Month", "MonthNum", "AIC", "BIC", "MAPE", "dAIC", "dBIC", "dMAPE", "PValue_new_var"]


def individual_month_deltas(df, target, base_cols, missing_months) -> pd.DataFrame:
    """AIC/BIC/MAPE for adding each missing month ALONE, for the 'ranked by
    impact' display (independent of whether it made the greedy cut)."""
    if not missing_months:
        return pd.DataFrame(columns=SOLO_DELTA_COLUMNS)
    work_df = df.copy()
    baseline = ols_fit(work_df, target, base_cols)
    rows = []
    for mm in missing_months:
        col = f"_solo_{MONTH_ABBR[mm - 1]}"
        work_df[col] = (work_df["Month"] == mm).astype(float)
        fit = ols_fit(work_df, target, base_cols + [col])
        rows.append({
            "Month": MONTH_ABBR[mm - 1],
            "MonthNum": mm,
            "AIC": fit.aic,
            "BIC": fit.bic,
            "MAPE": fit.mape,
            "dAIC": fit.aic - baseline.aic,
            "dBIC": fit.bic - baseline.bic,
            "dMAPE": fit.mape - baseline.mape,
            "PValue_new_var": fit.pvalue[-1] if len(fit.pvalue) else np.nan,
        })
    return pd.DataFrame(rows, columns=SOLO_DELTA_COLUMNS).sort_values("dAIC")


def build_recommendations(pw: ParsedWorkbook, p_thresh: float = 0.05, bic_pref: bool = True):
    """Top-level entry point: returns (recommendations: list[Recommendation],
    extras: dict of supporting tables for the UI)."""
    df = pw.data_df
    target = pw.target_name
    base_cols = list(pw.regressor_cols)

    removal_df = existing_binary_removal_candidates(pw.coef_df, p_thresh)
    trimmed_cols = list(base_cols)
    for _, row in removal_df.iterrows():
        var = str(row["Variable"])
        matches = [c for c in trimmed_cols if c.lower() in var.lower() or var.lower().endswith(c.lower())]
        for mvar in matches:
            if mvar in trimmed_cols:
                trimmed_cols.remove(mvar)

    accepted_months, trace_df = forward_select_months(
        df, target, trimmed_cols, pw.missing_months, bic_pref=bic_pref
    )
    solo_deltas = individual_month_deltas(df, target, trimmed_cols, pw.missing_months)

    continuous_cols = continuous_regressors(df, base_cols)
    sign_flags = sign_check(pw.coef_df, continuous_cols, p_thresh)

    month_bias = by_month_bias(pw.err_df)
    outliers = find_outliers(pw.err_df)

    recs = []
    rank = 1

    accepted_names = {MONTH_ABBR[mm - 1] for mm in accepted_months}
    for _, row in solo_deltas.iterrows():
        mon = row["Month"]
        is_yes = mon in accepted_names
        recs.append(Recommendation(
            rank=rank,
            category="Add month",
            variable=mon,
            add_yes_no="Yes" if is_yes else "No",
            delta_aic=row["dAIC"],
            delta_bic=row["dBIC"],
            delta_mape=row["dMAPE"],
            pvalue=row["PValue_new_var"],
            rationale=(
                "Part of the BIC-parsimonious accepted set (forward selection)."
                if is_yes else
                "Improves AIC alone but not part of the best BIC-parsimonious combination found; "
                "adding it would cost more in parameters than it buys in fit."
            ),
        ))
        rank += 1

    for _, row in removal_df.iterrows():
        recs.append(Recommendation(
            rank=rank,
            category="Remove existing",
            variable=str(row["Variable"]),
            add_yes_no="No",
            delta_aic=None,
            delta_bic=None,
            delta_mape=None,
            pvalue=row["PValue"],
            rationale=f"Not significant in the reported MetrixND fit (p={row['PValue']:.3f} > {p_thresh}).",
        ))
        rank += 1

    for flag in sign_flags:
        recs.append(Recommendation(
            rank=rank,
            category="Investigate",
            variable=flag.split(":")[0],
            add_yes_no="No",
            delta_aic=None, delta_bic=None, delta_mape=None, pvalue=None,
            rationale=flag,
        ))
        rank += 1

    recs.sort(key=lambda r: (r.delta_aic if r.delta_aic is not None else 0))
    for i, r in enumerate(recs, start=1):
        r.rank = i

    extras = {
        "removal_candidates": removal_df,
        "forward_trace": trace_df,
        "solo_deltas": solo_deltas,
        "month_bias": month_bias,
        "outliers": outliers,
        "sign_flags": sign_flags,
        "trimmed_cols": trimmed_cols,
        "accepted_months": [MONTH_ABBR[mm - 1] for mm in accepted_months],
    }
    return recs, extras
